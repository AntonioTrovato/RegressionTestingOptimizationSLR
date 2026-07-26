#!/usr/bin/env python3
"""
Imports quality appraisal results from the old Quality_Appraisal.xlsx into
the new QualityAppraisal.bib, for papers that are present in both (matched
by normalized title + author overlap).

Requires: openpyxl (not in the standard library)
    pip install openpyxl

Usage:
    python import_old_quality_appraisal.py [path_to_6_Quality_Appraisal_folder]

If no argument is given, the current directory is used
(assumed to be 6_Quality_Appraisal).

Expected input:
    6_Quality_Appraisal/QualityAppraisal.bib
    6_Quality_Appraisal/Quality_Appraisal.xlsx

Output:
    6_Quality_Appraisal/QualityAppraisal_WithOldScores.bib

For every paper found in BOTH files (matched by title + author), this adds:
    idq
    cdq_1_a, cdq_1_b, cdq_2_a, cdq_2_b, cdq_3_a, cdq_3_b,
    cdq_4_a, cdq_4_b, cdq_5_a, cdq_5_b, cdq_6_a, cdq_6_b   (only the ones
                                                             actually present
                                                             in the xlsx - PS
                                                             papers typically
                                                             only have 1-4)
    cdq_a
    cdq_b
    cdq
    fdq
Papers not found in the old xlsx are left untouched (they'll be appraised
from scratch - handled in a future step, as agreed).
"""

import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("This script requires the 'openpyxl' package.")
    print("Install it with:  pip install openpyxl")
    sys.exit(1)

# Column headers in the xlsx we care about (besides Author/Title, used for matching)
CDQ_FIELDS = []
for i in range(1, 7):
    CDQ_FIELDS.append(f"CDQ_{i}_A")
    CDQ_FIELDS.append(f"CDQ_{i}_B")
SCORE_FIELDS = ["IDQ"] + CDQ_FIELDS + ["CDQ_A", "CDQ_B", "CDQ", "FDQ"]


# --------------------------------------------------------------------------
# BibTeX parsing helpers
# --------------------------------------------------------------------------

def split_entries(text):
    starts = [m.start() for m in re.finditer(r'@\w+\s*\{', text)]
    entries = []
    for i, start in enumerate(starts):
        end = starts[i + 1] if i + 1 < len(starts) else len(text)
        entry = text[start:end].strip()
        if entry:
            entries.append(entry)
    return entries


def extract_field(entry, field):
    pattern = re.compile(r'\b' + re.escape(field) + r'\s*=\s*(\{|")', re.IGNORECASE)
    m = pattern.search(entry)
    if not m:
        return None

    open_char = m.group(1)
    start = m.end()

    if open_char == '"':
        end = entry.find('"', start)
        return entry[start:end].strip() if end != -1 else None

    depth = 1
    i = start
    while i < len(entry) and depth > 0:
        if entry[i] == '{':
            depth += 1
        elif entry[i] == '}':
            depth -= 1
        i += 1
    return entry[start:i - 1].strip()


def normalize_title(title):
    if not title:
        return None
    t = title.lower()
    t = re.sub(r'[{}\\]', '', t)
    t = re.sub(r'[^a-z0-9]+', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def author_last_names(author_field):
    """Extract a set of normalized 'name words' (>= 3 chars, to avoid
    matching on initials like 'a' or 'c') from an author field.

    We deliberately do NOT rely on a fixed format (e.g. BibTeX-style
    'Last, First and Last, First') because the old xlsx uses a different,
    inconsistent format (plain 'First Last', sometimes several authors
    concatenated with no separators at all, e.g.
    'ahmadreza saboor yaraghi mojtaba bagherzadeh nafiseh kahani lionel c briand').

    Instead we just tokenize into words and compare word-sets for overlap:
    this reliably catches shared surnames regardless of formatting."""
    if not author_field:
        return set()
    text = author_field.lower()
    text = re.sub(r'[{}\\]', ' ', text)
    words = re.findall(r'[a-z0-9]+', text)
    return {w for w in words if len(w) >= 3}


def add_fields(entry, fields_to_add):
    """Return a copy of the entry text with the given fields (list of
    (name, value) tuples) added just before the final closing brace."""
    entry = entry.rstrip()
    if not entry.endswith("}"):
        insertion = "".join(f",\n{name} = {{{value}}}" for name, value in fields_to_add)
        return entry + insertion + "\n}"

    body = entry[:-1].rstrip()
    if body.endswith(","):
        body = body[:-1]

    insertion = "".join(f",\n{name} = {{{value}}}" for name, value in fields_to_add)
    return body + insertion + "\n}"


# --------------------------------------------------------------------------
# XLSX loading
# --------------------------------------------------------------------------

def load_old_scores(xlsx_path):
    """Returns a list of dicts, one per row, with keys:
    'title_norm', 'authors' (set of last names), and all SCORE_FIELDS."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {}
    for i, h in enumerate(header_row):
        if h in ("Author", "Title") or h in SCORE_FIELDS:
            # Only keep the FIRST occurrence of each relevant header name
            if h not in col_index:
                col_index[h] = i

    missing = [c for c in ["Author", "Title"] + SCORE_FIELDS if c not in col_index]
    if missing:
        print(f"WARNING: expected columns not found in xlsx: {missing}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[col_index["Title"]] if "Title" in col_index else None
        author = row[col_index["Author"]] if "Author" in col_index else None
        title_norm = normalize_title(title)
        if not title_norm:
            continue

        entry = {
            "title_norm": title_norm,
            "authors": author_last_names(author),
        }
        for field in SCORE_FIELDS:
            if field in col_index:
                entry[field] = row[col_index[field]]
            else:
                entry[field] = None
        rows.append(entry)

    return rows


def build_title_index(old_rows):
    """Map normalized title -> list of matching xlsx rows (usually 1)."""
    index = {}
    for row in old_rows:
        index.setdefault(row["title_norm"], []).append(row)
    return index


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else "."
    folder = os.path.abspath(folder)

    bib_path = os.path.join(folder, "QualityAppraisal.bib")
    xlsx_path = os.path.join(folder, "Quality_Appraisal.xlsx")
    output_path = os.path.join(folder, "QualityAppraisal_WithOldScores.bib")

    if not os.path.isfile(bib_path):
        print(f"File not found: {bib_path}")
        sys.exit(1)
    if not os.path.isfile(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    print(f"Bib file:  {bib_path}")
    print(f"Xlsx file: {xlsx_path}")
    print(f"Output:    {output_path}\n")

    with open(bib_path, "r", encoding="utf-8") as f:
        text = f.read()
    entries = split_entries(text)

    old_rows = load_old_scores(xlsx_path)
    title_index = build_title_index(old_rows)

    print(f"Bib entries: {len(entries)}")
    print(f"Old xlsx rows loaded: {len(old_rows)}\n")

    matched_entries = []
    matched_count = 0
    title_only_conflict = 0  # title matched but no author overlap -> not used
    unmatched_count = 0

    for e in entries:
        title_norm = normalize_title(extract_field(e, "title"))
        authors = author_last_names(extract_field(e, "author"))

        candidates = title_index.get(title_norm, [])

        match_row = None
        for cand in candidates:
            if not cand["authors"] or not authors:
                # Can't verify via authors, but title normalized match is
                # already quite strong -> accept if it's the only candidate
                if len(candidates) == 1:
                    match_row = cand
                    break
                continue
            if cand["authors"] & authors:  # any overlap in last names
                match_row = cand
                break

        if match_row is None and candidates:
            title_only_conflict += 1

        if match_row is None:
            matched_entries.append(e)
            unmatched_count += 1
            continue

        fields_to_add = []
        idq = match_row.get("IDQ")
        if idq is not None:
            fields_to_add.append(("idq", idq))

        for i in range(1, 7):
            for letter in ("A", "B"):
                key = f"CDQ_{i}_{letter}"
                value = match_row.get(key)
                if value is not None and value != "":
                    fields_to_add.append((f"cdq_{i}_{letter.lower()}", value))

        for key, name in (("CDQ_A", "cdq_a"), ("CDQ_B", "cdq_b"), ("CDQ", "cdq")):
            value = match_row.get(key)
            if value is not None and value != "":
                fields_to_add.append((name, value))

        fdq = match_row.get("FDQ")
        if fdq is not None:
            fields_to_add.append(("fdq", fdq))

        if fields_to_add:
            matched_entries.append(add_fields(e, fields_to_add))
            matched_count += 1
        else:
            matched_entries.append(e)
            unmatched_count += 1

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n\n".join(matched_entries))
        if matched_entries:
            f.write("\n")

    print("=== RESULTS ===")
    print(f"Papers matched with old quality appraisal scores: {matched_count}")
    print(f"Papers with matching title but NO author overlap (skipped, please check manually): {title_only_conflict}")
    print(f"Papers left untouched (not found in old xlsx): {unmatched_count}")
    print(f"\nSaved: {output_path}")


if __name__ == "__main__":
    main()